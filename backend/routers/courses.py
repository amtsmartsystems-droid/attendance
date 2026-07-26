"""
ملف: backend/routers/courses.py
الوصف: API كاملة لإدارة الدورات التدريبية مع نظام عزل البيانات
"""
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from supabase import Client
import os
import io
from dotenv import load_dotenv

load_dotenv()
from supabase import create_client

url: str = os.getenv("SUPABASE_URL")
key: str = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
supabase: Client = create_client(url, key)

router = APIRouter(prefix="/api/courses", tags=["Courses"])


# ── نماذج البيانات ─────────────────────────────────────────────────────────

class CreateCourseRequest(BaseModel):
    title: str
    trainer_name: Optional[str] = None
    description: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None

class UpdateCourseRequest(BaseModel):
    title: Optional[str] = None
    trainer_name: Optional[str] = None
    description: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    is_active: Optional[bool] = None

class AddDoorRequest(BaseModel):
    door_name: str
    nfc_uid: str

@router.post("/{course_id}/add-door")
def add_new_door(course_id: str, req: AddDoorRequest):
    # أولاً: هل البطاقة موجودة مسبقاً في النظام؟
    existing = supabase.table("doors").select("*").eq("nfc_uid", req.nfc_uid).execute()
    
    if existing.data:
        # البطاقة موجودة — نحدّث اسمها ونربطها بالدورة
        res = supabase.table("doors").update({
            "door_name": req.door_name,
            "course_id": course_id
        }).eq("nfc_uid", req.nfc_uid).execute()
        return {"message": "تم تحديث البوابة وربطها بالدورة بنجاح", "door": res.data[0]}
    else:
        # بطاقة جديدة — نُدرجها مع door_id تلقائي
        import uuid as uuid_lib
        auto_door_id = f"DOOR_{req.nfc_uid.replace(':', '').upper()[:8]}"
        try:
            res = supabase.table("doors").insert({
                "door_id": auto_door_id,
                "door_name": req.door_name,
                "nfc_uid": req.nfc_uid,
                "course_id": course_id,
                "is_active": True
            }).execute()
            return {"message": "تم إضافة البوابة الجديدة وربطها بنجاح", "door": res.data[0]}
        except Exception as e:
            raise HTTPException(status_code=400, detail=str(e))

class AssignDoorRequest(BaseModel):
    door_id: str   # UUID الباب

class EnrollTraineeRequest(BaseModel):
    employee_id: str  # UUID الموظف/المتدرب


# ── GET /api/courses — قائمة كل الدورات مع إحصاءات ────────────────────────

@router.get("/")
def get_all_courses():
    res = supabase.table("courses").select("*").order("created_at", desc=True).execute()
    courses = res.data or []

    enriched = []
    for course in courses:
        # عدد المتدربين المسجلين
        enroll_res = supabase.table("course_enrollments") \
            .select("id", count="exact") \
            .eq("course_id", course["id"]) \
            .execute()
        trainee_count = enroll_res.count or 0

        # عدد سجلات الحضور اليوم
        import datetime
        today = datetime.date.today().isoformat()
        att_res = supabase.table("attendance_logs") \
            .select("id", count="exact") \
            .eq("course_id", course["id"]) \
            .gte("recorded_at", f"{today}T00:00:00+03:00") \
            .execute()
        today_attendance = att_res.count or 0

        enriched.append({
            **course,
            "trainee_count": trainee_count,
            "today_attendance": today_attendance,
        })

    return {"courses": enriched}


# ── GET /api/courses/{id} — تفاصيل دورة واحدة ────────────────────────────

@router.get("/{course_id}")
def get_course(course_id: str):
    res = supabase.table("courses").select("*").eq("id", course_id).execute()
    if not res.data:
        raise HTTPException(status_code=404, detail="الدورة غير موجودة")
    course = res.data[0]

    # المتدربون المسجلون
    enroll_res = supabase.table("course_enrollments") \
        .select("*, employees(id, emp_id, full_name, phone_number, status, is_device_bound)") \
        .eq("course_id", course_id) \
        .execute()

    # الأبواب المرتبطة
    doors_res = supabase.table("doors").select("*").eq("course_id", course_id).execute()

    return {
        "course": course,
        "trainees": [e["employees"] for e in (enroll_res.data or []) if e.get("employees")],
        "doors": doors_res.data or [],
    }


# ── POST /api/courses — إنشاء دورة جديدة ─────────────────────────────────

@router.post("/")
def create_course(req: CreateCourseRequest):
    import datetime, uuid as uuid_lib
    course_code = f"COURSE_{datetime.date.today().year}_{str(uuid_lib.uuid4())[:6].upper()}"

    res = supabase.table("courses").insert({
        "course_code": course_code,
        "title": req.title,
        "trainer_name": req.trainer_name,
        "description": req.description,
        "start_date": req.start_date,
        "end_date": req.end_date,
        "is_active": True,
    }).execute()

    if not res.data:
        raise HTTPException(status_code=500, detail="فشل في إنشاء الدورة")

    return {"message": "تم إنشاء الدورة بنجاح", "course": res.data[0]}


# ── PUT /api/courses/{id} — تعديل دورة ───────────────────────────────────

@router.put("/{course_id}")
def update_course(course_id: str, req: UpdateCourseRequest):
    update_data = {k: v for k, v in req.dict().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="لا توجد بيانات للتحديث")

    res = supabase.table("courses").update(update_data).eq("id", course_id).execute()
    if not res.data:
        raise HTTPException(status_code=404, detail="الدورة غير موجودة")

    return {"message": "تم تحديث الدورة", "course": res.data[0]}


# ── DELETE /api/courses/{id} — حذف دورة ──────────────────────────────────

@router.delete("/{course_id}")
def delete_course(course_id: str):
    res = supabase.table("courses").delete().eq("id", course_id).execute()
    return {"message": "تم حذف الدورة"}


# ── POST /api/courses/{id}/assign-door — ربط باب NFC بالدورة ─────────────

@router.post("/{course_id}/assign-door")
def assign_door_to_course(course_id: str, req: AssignDoorRequest):
    # التحقق من وجود الدورة
    course_res = supabase.table("courses").select("id").eq("id", course_id).execute()
    if not course_res.data:
        raise HTTPException(status_code=404, detail="الدورة غير موجودة")

    # ربط الباب بالدورة
    res = supabase.table("doors").update({"course_id": course_id}).eq("id", req.door_id).execute()
    if not res.data:
        raise HTTPException(status_code=404, detail="الباب غير موجود")

    return {"message": "تم ربط الباب بالدورة بنجاح", "door": res.data[0]}


# ── POST /api/courses/{id}/enroll — تسجيل متدرب في دورة ─────────────────

@router.post("/{course_id}/enroll")
def enroll_trainee(course_id: str, req: EnrollTraineeRequest):
    try:
        res = supabase.table("course_enrollments").insert({
            "course_id": course_id,
            "employee_id": req.employee_id,
        }).execute()
        return {"message": "تم تسجيل المتدرب في الدورة"}
    except Exception as e:
        if "unique" in str(e).lower():
            raise HTTPException(status_code=409, detail="المتدرب مسجل مسبقاً في هذه الدورة")
        raise HTTPException(status_code=500, detail=str(e))


# ── GET /api/courses/{id}/attendance — سجلات الحضور للدورة ──────────────

@router.get("/{course_id}/attendance")
def get_course_attendance(course_id: str, date: Optional[str] = None):
    query = supabase.table("attendance_logs") \
        .select("*, employees(full_name, emp_id, phone_number), doors(door_name)") \
        .eq("course_id", course_id) \
        .order("recorded_at", desc=True)

    if date:
        query = query.gte("recorded_at", f"{date}T00:00:00+03:00") \
                     .lte("recorded_at", f"{date}T23:59:59+03:00")

    res = query.execute()
    return {"attendance": res.data or [], "total": len(res.data or [])}


# ── GET /api/courses/{id}/export — تصدير Excel ───────────────────────────

@router.get("/{course_id}/export")
def export_course_attendance(course_id: str):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import datetime
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl غير مثبت على السيرفر")

    # جلب بيانات الدورة
    course_res = supabase.table("courses").select("*").eq("id", course_id).execute()
    if not course_res.data:
        raise HTTPException(status_code=404, detail="الدورة غير موجودة")
    course = course_res.data[0]

    # جلب سجلات الحضور
    att_res = supabase.table("attendance_logs") \
        .select("*, employees(full_name, emp_id, phone_number, department), doors(door_name)") \
        .eq("course_id", course_id) \
        .order("recorded_at", desc=True) \
        .execute()

    records = att_res.data or []

    # إنشاء ملف Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "سجل الحضور"
    ws.sheet_view.rightToLeft = True

    # ألوان وتنسيقات
    header_fill = PatternFill(start_color="1A2B4A", end_color="1A2B4A", fill_type="solid")
    accent_fill = PatternFill(start_color="00B4D8", end_color="00B4D8", fill_type="solid")
    alt_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    title_font = Font(name="Arial", bold=True, color="FFFFFF", size=14)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # ─ عنوان التقرير ─
    ws.merge_cells("A1:G1")
    ws["A1"] = f"تقرير حضور — {course['title']}"
    ws["A1"].font = title_font
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    ws["A2"] = f"المدرب: {course.get('trainer_name', '—')}   |   تاريخ التصدير: {datetime.date.today()}"
    ws["A2"].font = Font(name="Arial", size=10, color="00B4D8", bold=True)
    ws["A2"].fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
    ws["A2"].alignment = Alignment(horizontal="center")

    # ─ رؤوس الأعمدة ─
    headers = ["#", "رقم المتدرب", "الاسم الكامل", "القسم/الدورة", "نوع الحركة", "التاريخ", "الوقت"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = accent_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[3].height = 25

    # ─ البيانات ─
    for row_idx, record in enumerate(records, 1):
        emp = record.get("employees") or {}
        door = record.get("doors") or {}

        import pytz
        from datetime import datetime as dt_cls
        recorded_at_str = record.get("recorded_at", "")
        try:
            amman_tz = pytz.timezone("Asia/Amman")
            recorded_dt = dt_cls.fromisoformat(recorded_at_str.replace("Z", "+00:00"))
            recorded_local = recorded_dt.astimezone(amman_tz)
            date_str = recorded_local.strftime("%Y-%m-%d")
            time_str = recorded_local.strftime("%I:%M %p")
        except Exception:
            date_str = recorded_at_str[:10] if recorded_at_str else "—"
            time_str = recorded_at_str[11:16] if len(recorded_at_str) > 16 else "—"

        row_data = [
            row_idx,
            emp.get("emp_id", "—"),
            emp.get("full_name", "—"),
            emp.get("department", course.get("title", "—")),
            record.get("movement_type", "—"),
            date_str,
            time_str,
        ]

        fill = alt_fill if row_idx % 2 == 0 else None
        excel_row = row_idx + 3
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=excel_row, column=col, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            if fill:
                cell.fill = fill

    # ─ ضبط أعرض الأعمدة ─
    col_widths = [6, 16, 25, 20, 14, 14, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    # ─ إرجاع الملف ─
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"attendance_{course.get('course_code', course_id)}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

