from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from supabase import create_client, Client
from pydantic import BaseModel
from dotenv import load_dotenv
import pandas as pd
import os
import io
import random
import string
from datetime import datetime, timezone, timedelta

# ── تحميل .env ──────────────────────────────────────────
load_dotenv()  # يقرأ من ملف .env تلقائياً

# ============================================================
# مكتبات التقارير
# ============================================================
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# استيراد المسارات الجديدة (Routers)
from routers import schedules, leaves, onboarding, courses

# ============================================================
class EmployeeCreate(BaseModel):
    full_name: str
    department: str

app = FastAPI(title="AMT Attendance Admin API")

# ── CORS ────────────────────────────────────────────────
CORS_ORIGINS = os.getenv("CORS_ORIGINS", "http://localhost:8080").split(",")
ENVIRONMENT = os.getenv("ENVIRONMENT", "development")

# في وضع التطوير: اسمح بكل الأصل | الإنتاج: فقط النطاقات المحددة
allowed_origins = ["*"] if ENVIRONMENT == "development" else CORS_ORIGINS

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# تسجيل المسارات (Routers)
app.include_router(schedules.router)
app.include_router(leaves.router)
app.include_router(onboarding.router)
app.include_router(courses.router)

# Keep-Alive endpoint لمنع نوم السيرفر على Render
@app.get("/ping")
def ping():
    return {"status": "alive", "message": "AMT Attendance Server is running"}

# ── Supabase ─────────────────────────────────────────────
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")

if not SUPABASE_URL or not SUPABASE_KEY:
    raise RuntimeError(
        "❌ خطأ حرج: SUPABASE_URL أو SUPABASE_SERVICE_ROLE_KEY غير محددين!\n"
        "تأكد من وجود ملف .env أو متغيرات البيئة في السيرفر السحابي."
    )

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
HMAC_SECRET = os.getenv("HMAC_SECRET", "")
RIYADH_TZ = timezone(timedelta(hours=3))

# ============================================================
# مساعد: تحويل التوقيت لمنطقة الرياض
# ============================================================
def to_riyadh(dt_str: str) -> str:
    try:
        dt = datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
        local = dt.astimezone(RIYADH_TZ)
        return local.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return dt_str


# ============================================================
# GET /api/stats
# ============================================================
@app.get("/api/stats")
def get_dashboard_stats():
    try:
        employees_res = supabase.table("employees").select("id", count="exact").execute()
        doors_res = supabase.table("doors").select("id", count="exact").execute()

        today = datetime.now(RIYADH_TZ).date().isoformat()
        attendance_res = (
            supabase.table("attendance_logs")
            .select("id", count="exact")
            .gte("recorded_at", f"{today}T00:00:00+03:00")
            .execute()
        )

        return {
            "success": True,
            "data": {
                "total_employees": employees_res.count or 0,
                "total_doors": doors_res.count or 0,
                "today_attendance": attendance_res.count or 0,
            },
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================
# GET /api/attendance  — آخر 100 سجل (للداشبورد القديم)
# ============================================================
@app.get("/api/attendance")
def get_attendance_logs():
    try:
        res = (
            supabase.table("attendance_logs")
            .select("*, employees(full_name, emp_id), doors(door_name)")
            .order("recorded_at", desc=True)
            .limit(100)
            .execute()
        )
        return {"success": True, "data": res.data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================
# GET /api/attendance/all — كل السجلات (للشاشة الجديدة)
# ============================================================
@app.get("/api/attendance/all")
def get_all_attendance_logs():
    try:
        res = (
            supabase.table("attendance_logs")
            .select("*, employees(full_name, emp_id, department), doors(door_name)")
            .order("recorded_at", desc=True)
            .execute()
        )
        return {"success": True, "data": res.data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================
# GET /api/attendance/live — آخر 10 دقائق فقط
# ============================================================
@app.get("/api/attendance/live")
def get_live_attendance():
    try:
        cutoff = (datetime.now(timezone.utc) - timedelta(minutes=10)).isoformat()
        res = (
            supabase.table("attendance_logs")
            .select("*, employees(full_name, emp_id, department), doors(door_name)")
            .gte("recorded_at", cutoff)
            .order("recorded_at", desc=True)
            .execute()
        )
        return {"success": True, "data": res.data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================
# GET /api/employees
# ============================================================
@app.get("/api/employees")
def get_employees():
    try:
        res = supabase.table("employees").select("*").order("created_at", desc=True).execute()
        return {"success": True, "data": res.data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================
# POST /api/employees
# ============================================================
@app.post("/api/employees")
def create_employee(emp: EmployeeCreate):
    try:
        pin = "".join(random.choices(string.digits, k=6))
        emp_id = f"EMP_{''.join(random.choices(string.digits, k=4))}"
        new_emp = {
            "emp_id": emp_id,
            "full_name": emp.full_name,
            "department": emp.department,
            "pin_code": pin,
            "is_active": True,
            "is_device_bound": False,
        }
        res = supabase.table("employees").insert(new_emp).execute()
        return {
            "success": True,
            "data": res.data[0],
            "pin": pin,
            "message": "تم إضافة الموظف بنجاح",
        }
    except Exception as e:
        print(f"Create Employee Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================
# DELETE /api/employees/{id}
# ============================================================
@app.delete("/api/employees/{emp_uuid}")
def delete_employee(emp_uuid: str):
    try:
        supabase.table("employees").delete().eq("id", emp_uuid).execute()
        return {"success": True, "message": "تم حذف الموظف بنجاح"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================
# POST /api/employees/{id}/reset-device — إعادة ضبط جهاز الموظف
# ============================================================
@app.post("/api/employees/{emp_uuid}/reset-device")
def reset_employee_device(emp_uuid: str):
    """
    يُلغي ربط جهاز الموظف ويولد رمز PIN جديد.
    يُستخدم عند تغيير الهاتف أو فقدانه.
    """
    try:
        new_pin = "".join(random.choices(string.digits, k=6))
        res = supabase.table("employees").update({
            "is_device_bound": False,
            "device_id": None,
            "pin_code": new_pin,
        }).eq("id", emp_uuid).execute()

        if not res.data:
            raise HTTPException(status_code=404, detail="الموظف غير موجود")

        emp = res.data[0]
        return {
            "success": True,
            "new_pin": new_pin,
            "emp_id": emp.get("emp_id"),
            "message": "تم إعادة ضبط الجهاز بنجاح — أرسل الـ PIN الجديد للموظف",
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Reset Device Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class LoginRequest(BaseModel):
    pin_code: str
    device_id: str

# ============================================================
# POST /api/employees/login — تسجيل الدخول وربط الجهاز
# ============================================================
@app.post("/api/employees/login")
def login_employee(req: LoginRequest):
    """
    يُستخدم من تطبيق الموبايل لتسجيل الدخول برمز PIN.
    يتحقق من الرمز ويربط الجهاز بالموظف.
    """
    try:
        # 1. البحث عن الموظف
        res = supabase.table("employees").select("*").eq("pin_code", req.pin_code).execute()
        if not res.data:
            raise HTTPException(status_code=401, detail="الرمز السري غير صحيح")
        
        emp = res.data[0]
        
        # 2. التحقق من حالة الربط
        if emp.get("is_device_bound"):
            raise HTTPException(status_code=403, detail="هذا الرمز السري مستخدم مسبقاً على جهاز آخر")
            
        # 3. ربط الجهاز
        update_res = supabase.table("employees").update({
            "is_device_bound": True,
            "device_id": req.device_id
        }).eq("id", emp["id"]).execute()
        
        if not update_res.data:
            raise HTTPException(status_code=500, detail="فشل في ربط الجهاز")
            
        return {
            "success": True,
            "emp_id": emp["emp_id"],
            "full_name": emp["full_name"],
            "message": "تم تسجيل الدخول وربط الجهاز بنجاح"
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Login Error: {e}")
        raise HTTPException(status_code=500, detail="حدث خطأ في الاتصال بقاعدة البيانات")

# ============================================================
# GET /api/reports/excel — تقرير Excel احترافي متعدد الأوراق
# ============================================================
@app.get("/api/reports/excel")
def export_attendance_excel():
    try:
        res = (
            supabase.table("attendance_logs")
            .select("*, employees(full_name, emp_id, department), doors(door_name, location)")
            .order("recorded_at", desc=True)
            .execute()
        )
        if not res.data:
            raise HTTPException(status_code=404, detail="لا توجد بيانات")

        wb = openpyxl.Workbook()

        # ── الألوان ──────────────────────────────────────
        COLOR_HEADER   = "0D1A2D"
        COLOR_ACCENT   = "00E5FF"
        COLOR_ROW_ODD  = "162032"
        COLOR_ROW_EVEN = "1A2842"
        COLOR_GREEN    = "1B5E20"
        COLOR_ORANGE   = "E65100"

        thin = Side(border_style="thin", color="2C3E50")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def style_header(ws, row, columns):
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=row, column=col_idx, value=col_name)
                cell.font = Font(bold=True, color=COLOR_ACCENT, size=11)
                cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

        # ─────────────────────────────────────────────────
        # ورقة 1: جميع السجلات
        # ─────────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "جميع السجلات"
        ws1.sheet_view.rightToLeft = True
        ws1.row_dimensions[1].height = 20

        cols1 = ["التاريخ والوقت", "رقم الموظف", "اسم الموظف", "القسم", "البوابة", "نوع الحركة", "مصدر البيانات"]
        style_header(ws1, 1, cols1)

        for i, row in enumerate(res.data, start=2):
            emp = row.get("employees") or {}
            door = row.get("doors") or {}
            mtype = row.get("movement_type", "")
            values = [
                to_riyadh(row.get("recorded_at", "")),
                emp.get("emp_id", "-"),
                emp.get("full_name", "غير معروف"),
                emp.get("department", "-"),
                door.get("door_name", "غير معروف"),
                mtype,
                "أوفلاين" if row.get("is_offline_sync") else "أونلاين",
            ]
            fill_color = COLOR_ROW_ODD if i % 2 == 0 else COLOR_ROW_EVEN
            for col_idx, val in enumerate(values, start=1):
                cell = ws1.cell(row=i, column=col_idx, value=val)
                cell.font = Font(color="FFFFFF", size=10)
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
                if col_idx == 6:  # نوع الحركة
                    cell.font = Font(
                        bold=True,
                        color=COLOR_GREEN if mtype == "دخول" else COLOR_ORANGE,
                        size=10,
                    )

        col_widths1 = [22, 14, 22, 16, 18, 12, 14]
        for i, w in enumerate(col_widths1, start=1):
            ws1.column_dimensions[get_column_letter(i)].width = w

        # ─────────────────────────────────────────────────
        # ورقة 2: ملخص يومي
        # ─────────────────────────────────────────────────
        ws2 = wb.create_sheet("الملخص اليومي")
        ws2.sheet_view.rightToLeft = True

        daily: dict = {}
        for row in res.data:
            emp = row.get("employees") or {}
            mtype = row.get("movement_type", "")
            try:
                dt = datetime.fromisoformat(row["recorded_at"].replace("Z", "+00:00")).astimezone(RIYADH_TZ)
                day_key = dt.date().isoformat()
            except Exception:
                continue
            if day_key not in daily:
                daily[day_key] = {"دخول": 0, "خروج": 0, "موظفون": set()}
            daily[day_key][mtype] = daily[day_key].get(mtype, 0) + 1
            emp_id = emp.get("emp_id")
            if emp_id:
                daily[day_key]["موظفون"].add(emp_id)

        cols2 = ["التاريخ", "عدد حركات الدخول", "عدد حركات الخروج", "عدد الموظفين الحاضرين"]
        style_header(ws2, 1, cols2)
        for i, (day, data) in enumerate(sorted(daily.items(), reverse=True), start=2):
            fill_color = COLOR_ROW_ODD if i % 2 == 0 else COLOR_ROW_EVEN
            values = [day, data.get("دخول", 0), data.get("خروج", 0), len(data["موظفون"])]
            for col_idx, val in enumerate(values, start=1):
                cell = ws2.cell(row=i, column=col_idx, value=val)
                cell.font = Font(color="FFFFFF", size=10)
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

        for i, w in enumerate([16, 22, 22, 26], start=1):
            ws2.column_dimensions[get_column_letter(i)].width = w

        # ─────────────────────────────────────────────────
        # ورقة 3: ملخص حسب الموظف
        # ─────────────────────────────────────────────────
        ws3 = wb.create_sheet("ملخص حسب الموظف")
        ws3.sheet_view.rightToLeft = True

        emp_summary: dict = {}
        for row in res.data:
            emp = row.get("employees") or {}
            emp_id = emp.get("emp_id", "?")
            emp_name = emp.get("full_name", "غير معروف")
            dept = emp.get("department", "-")
            mtype = row.get("movement_type", "")
            if emp_id not in emp_summary:
                emp_summary[emp_id] = {"name": emp_name, "dept": dept, "دخول": 0, "خروج": 0}
            emp_summary[emp_id][mtype] = emp_summary[emp_id].get(mtype, 0) + 1

        cols3 = ["رقم الموظف", "الاسم", "القسم", "مجموع الدخول", "مجموع الخروج", "إجمالي الحركات"]
        style_header(ws3, 1, cols3)
        for i, (emp_id, data) in enumerate(emp_summary.items(), start=2):
            fill_color = COLOR_ROW_ODD if i % 2 == 0 else COLOR_ROW_EVEN
            total = data["دخول"] + data["خروج"]
            values = [emp_id, data["name"], data["dept"], data["دخول"], data["خروج"], total]
            for col_idx, val in enumerate(values, start=1):
                cell = ws3.cell(row=i, column=col_idx, value=val)
                cell.font = Font(color="FFFFFF", size=10)
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

        for i, w in enumerate([14, 22, 16, 16, 16, 18], start=1):
            ws3.column_dimensions[get_column_letter(i)].width = w

        # ── حفظ وإرجاع الملف ────────────────────────────
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        today_str = datetime.now(RIYADH_TZ).strftime("%Y-%m-%d")
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=AMT_Report_{today_str}.xlsx"},
        )
    except HTTPException:
        raise
    except Exception as e:
        print(f"Excel Export Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================
# GET /api/reports/pdf — تقرير PDF تنفيذي
# ============================================================
@app.get("/api/reports/pdf")
def export_attendance_pdf():
    try:
        res = (
            supabase.table("attendance_logs")
            .select("*, employees(full_name, emp_id, department), doors(door_name)")
            .order("recorded_at", desc=True)
            .execute()
        )

        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=1.5 * cm,
            leftMargin=1.5 * cm,
            topMargin=1.5 * cm,
            bottomMargin=1.5 * cm,
        )

        styles = getSampleStyleSheet()
        story = []

        # ── العنوان ───────────────────────────────────────
        title_style = ParagraphStyle(
            "title",
            parent=styles["Title"],
            fontSize=18,
            textColor=colors.HexColor("#00E5FF"),
            spaceAfter=4,
            alignment=1,
        )
        sub_style = ParagraphStyle(
            "sub",
            parent=styles["Normal"],
            fontSize=10,
            textColor=colors.HexColor("#AAAAAA"),
            spaceAfter=12,
            alignment=1,
        )

        story.append(Paragraph("تقرير الحضور والانصراف — AMT", title_style))
        today_str = datetime.now(RIYADH_TZ).strftime("%Y/%m/%d %H:%M")
        story.append(Paragraph(f"تاريخ الإصدار: {today_str}", sub_style))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#2C3E50")))
        story.append(Spacer(1, 0.3 * cm))

        # ── إحصائيات سريعة ────────────────────────────────
        total = len(res.data)
        entries = sum(1 for r in res.data if r.get("movement_type") == "دخول")
        exits = total - entries
        emp_set = {(r.get("employees") or {}).get("emp_id") for r in res.data if r.get("employees")}

        stat_data = [
            ["إجمالي الحركات", "حركات الدخول", "حركات الخروج", "عدد الموظفين"],
            [str(total), str(entries), str(exits), str(len(emp_set))],
        ]
        stat_table = Table(stat_data, colWidths=[6 * cm] * 4)
        stat_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0D1A2D")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#00E5FF")),
            ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#162032")),
            ("TEXTCOLOR", (0, 1), (-1, 1), colors.white),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("FONTSIZE", (0, 1), (-1, 1), 16),
            ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, 1), [colors.HexColor("#162032")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#2C3E50")),
            ("TOPPADDING", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ("ROUNDEDCORNERS", [6]),
        ]))
        story.append(stat_table)
        story.append(Spacer(1, 0.5 * cm))

        # ── جدول السجلات (آخر 50) ────────────────────────
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#2C3E50")))
        story.append(Spacer(1, 0.3 * cm))

        table_header = ["التاريخ والوقت", "رقم الموظف", "الاسم", "القسم", "البوابة", "النوع"]
        table_data = [table_header]

        for row in res.data[:50]:
            emp = row.get("employees") or {}
            door = row.get("doors") or {}
            table_data.append([
                to_riyadh(row.get("recorded_at", ""))[:16],
                emp.get("emp_id", "-"),
                emp.get("full_name", "غير معروف"),
                emp.get("department", "-"),
                door.get("door_name", "-"),
                row.get("movement_type", "-"),
            ])

        col_w = [4 * cm, 2.8 * cm, 4.5 * cm, 3 * cm, 3.5 * cm, 2.5 * cm]
        tbl = Table(table_data, colWidths=col_w, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0D1A2D")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#00E5FF")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("TEXTCOLOR", (0, 1), (-1, -1), colors.white),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.HexColor("#162032"), colors.HexColor("#1A2842")]),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#2C3E50")),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(tbl)

        if len(res.data) > 50:
            story.append(Spacer(1, 0.3 * cm))
            story.append(Paragraph(
                f"* يعرض الـ PDF آخر 50 سجلاً. لعرض جميع السجلات ({total})، قم بتصدير ملف Excel.",
                ParagraphStyle("note", fontSize=8, textColor=colors.HexColor("#AAAAAA"), alignment=1),
            ))

        doc.build(story)
        output.seek(0)

        today_str2 = datetime.now(RIYADH_TZ).strftime("%Y-%m-%d")
        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=AMT_Report_{today_str2}.pdf"},
        )
    except Exception as e:
        print(f"PDF Export Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)

# ============================================================
# POST /api/attendance/record � ????? ?????? ?? ?????? ?? ??????
# ============================================================
class AttendanceRecord(BaseModel):
    emp_id: str
    door_nfc_uid: str
    device_id: Optional[str] = None
    offline_timestamp: Optional[str] = None
    is_offline_sync: bool = False
    nonce: Optional[str] = None

@app.post('/api/attendance/record')
def record_attendance(req: AttendanceRecord):
    try:
        import uuid as uuid_lib
        emp_res = supabase.table('employees').select('*').eq('emp_id', req.emp_id).execute()
        if not emp_res.data:
            raise HTTPException(status_code=404, detail='?????? ??? ?????')
        emp = emp_res.data[0]
        if req.device_id and emp.get('is_device_bound'):
            reg_dev = emp.get('device_id', '')
            if reg_dev and reg_dev != req.device_id:
                raise HTTPException(status_code=403, detail='????? ??????: ?????? ?? ????? ?????? ??????. ????? ?? ???????.')
        if emp.get('status') == 'pending':
            raise HTTPException(status_code=403, detail='?????? ?? ?????? ???? ???')
        door_res = supabase.table('doors').select('*').eq('nfc_uid', req.door_nfc_uid).execute()
        if not door_res.data:
            raise HTTPException(status_code=404, detail='????? ??? ????')
        door = door_res.data[0]
        last_log = supabase.table('attendance_logs').select('movement_type').eq('employee_id', emp['id']).order('recorded_at', desc=True).limit(1).execute()
        last_type = last_log.data[0]['movement_type'] if last_log.data else '????'
        movement = '????' if last_type == '????' else '????'
        record_time = req.offline_timestamp if req.offline_timestamp else datetime.now(timezone.utc).isoformat()
        nonce_val = req.nonce or str(uuid_lib.uuid4())
        supabase.table('attendance_logs').insert({'employee_id': emp['id'], 'door_id': door['id'], 'recorded_at': record_time, 'movement_type': movement, 'nonce': nonce_val, 'device_info': {'device_id': req.device_id, 'is_offline': req.is_offline_sync}}).execute()
        return {'success': True, 'movement_type': movement, 'employee_name': emp['full_name'], 'door_name': door['door_name'], 'message': f'?? ????? {movement} ?????'}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))





from fastapi.staticfiles import StaticFiles
import os

admin_path = os.path.join(os.path.dirname(__file__), "admin_web")
if os.path.exists(admin_path):
    app.mount("/admin", StaticFiles(directory=admin_path, html=True), name="admin")
else:
    @app.get("/admin")
    def admin_not_found():
        return {"detail": "Admin Dashboard Not Found at " + admin_path}

